from PyPDF2 import PdfMerger, PdfReader, PdfWriter
import fitz
import os
from pdf2docx import Converter
import tabula
import pandas as pd
import shutil
import subprocess
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Image, Paragraph, Frame
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet
from pdf2image import convert_from_path
from fpdf import FPDF
import asyncio
from pyppeteer import launch
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
import argparse
import uuid


def merge_pdfs(pdf_list, output):
    merger = PdfMerger()
    for pdf in pdf_list:
        merger.append(pdf)
    merger.write(output)
    merger.close()


def split_pdf(input_pdf, start_page, end_page, output_pdf):
    reader = PdfReader(input_pdf)
    writer = PdfWriter()
    for page in range(start_page, end_page):
        writer.add_page(reader.pages[page])
    with open(output_pdf, 'wb') as output:
        writer.write(output)


def compress_pdf(input_pdf, output_pdf, dpi=100):
    doc = fitz.open(input_pdf)
    new_doc = fitz.open()

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        rect = page.rect
        pix = page.get_pixmap(dpi=dpi)
        unique_id = uuid.uuid4()
        image_path = f"/tmp/compress_{unique_id}_temp_page_{page_num}.jpg"
        pix.save(image_path)
        new_page = new_doc.new_page(width=rect.width, height=rect.height)
        new_page.insert_image(rect, filename=image_path)
        os.remove(image_path)

    new_doc.save(output_pdf, garbage=4, deflate=True)
    new_doc.close()
    doc.close()


def pdf_to_word(input_pdf, output_docx):
    cv = Converter(input_pdf)
    cv.convert(output_docx, start=0, end=None)
    cv.close()


def pdf_to_excel(input_pdf, output_excel):
    options = {
        'pages': 'all',
        'multiple_tables': True,
        'stream': True,
    }
    dfs = tabula.read_pdf(input_pdf, **options)
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=f'Sheet{i + 1}', index=False)


def word_to_pdf(input_docx, output_pdf):
    output_dir = os.path.dirname(output_pdf)
    os.makedirs(output_dir, exist_ok=True)

    libreOfficePath = "libreoffice"
    commandStrings = [libreOfficePath, "--headless", "--convert-to", "pdf", "--outdir", output_dir, input_docx]
    retCode = subprocess.call(commandStrings)

    if retCode == 0:
        temp_pdf_path = os.path.join(output_dir, os.path.splitext(os.path.basename(input_docx))[0] + ".pdf")
        shutil.move(temp_pdf_path, output_pdf)
    else:
        print(f"Looks like there is an error in pdf conversion process with return code {retCode}")


def excel_to_pdf(input_xlsx, output_pdf):
    wb = load_workbook(input_xlsx)
    pdf = SimpleDocTemplate(output_pdf, pagesize=letter)
    pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf', 'utf-8'))
    styles = getSampleStyleSheet()
    style_paragraph = styles["Normal"]
    style_paragraph.fontName = "DejaVuSans"
    elements = []
    for ws in wb:
        max_row = ws.max_row
        max_col = ws.max_column
        data = []
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)
                    cell_value = Paragraph(cell_value, style_paragraph)
                row_data.append(cell_value if cell_value else "")
            data.append(row_data)

        table = Table(data)
        elements.append(table)
        elements.append(PageBreak())

    pdf.build(elements)


def pdf_to_jpg(input_pdf, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    images = convert_from_path(input_pdf)
    for i, image in enumerate(images):
        output_path = os.path.join(output_folder, f'page_{i + 1}.jpg')
        image.save(output_path, 'JPEG')
    shutil.make_archive(output_folder, 'zip', output_folder)


def custom_sort_key(filename):
    name, extension = os.path.splitext(filename)
    parts = name.split('_')
    if len(parts) > 1 and parts[1].isdigit():
        return int(parts[1])
    else:
        return filename


def jpg_to_pdf(input_folder, output_pdf):
    pdf = FPDF()
    filenames = os.listdir(input_folder)
    sorted_filenames = sorted(filenames, key=custom_sort_key)
    for filename in sorted_filenames:
        if filename.endswith(".jpg"):
            pdf.add_page()
            image_path = os.path.join(input_folder, filename)
            pdf.image(image_path, 0, 0, 210, 297)
    pdf.output(output_pdf, "F")


def create_page_pdf(num, tmp, position=(105*mm, 20*mm)):
    c = canvas.Canvas(tmp)
    for i in range(1, num + 1):
        c.drawString(position[0], position[1], str(i))
        c.showPage()
    c.save()


def add_page_numbers(pdf_path, newpath, position=(100, 20)):
    unique_id = uuid.uuid4()
    tmp = f"/tmp/add_page_numbers{unique_id}_temp.pdf"
    position = (position[0]*mm, position[1]*mm)
    writer = PdfWriter()
    with open(pdf_path, "rb") as f:
        reader = PdfReader(f)
        n = len(reader.pages)
        create_page_pdf(n, tmp, position)

        with open(tmp, "rb") as ftmp:
            number_pdf = PdfReader(ftmp)
            for p in range(n):
                page = reader.pages[p]
                number_layer = number_pdf.pages[p]
                page.merge_page(number_layer)
                writer.add_page(page)
            if len(writer.pages) > 0:
                with open(newpath, "wb") as f:
                    writer.write(f)
        os.remove(tmp)


def protect_pdf(input_pdf, output_pdf, password):
    reader = PdfReader(input_pdf)
    writer = PdfWriter()

    for page_num in range(len(reader.pages)):
        writer.add_page(reader.pages[page_num])

    writer.encrypt(password)

    with open(output_pdf, "wb") as output_file:
        writer.write(output_file)


def unlock_pdf(input_pdf, output_pdf, password):
    reader = PdfReader(input_pdf)
    if reader.is_encrypted:
        reader.decrypt(password)

    writer = PdfWriter()

    for page_num in range(len(reader.pages)):
        writer.add_page(reader.pages[page_num])

    with open(output_pdf, "wb") as output_file:
        writer.write(output_file)


if __name__ == '__main__':
    #pdf_file = 'test.pdf'
    #merge_pdfs([pdf_file for _ in range(100)], 'media/merge_pdfs.pdf')
    #split_pdf('media/merge_pdfs.pdf', 15, 42, 'media/split.pdf')
    #compress_pdf(pdf_file, 'media/compress_pdf.pdf', 100)
    #pdf_to_word(pdf_file, 'media/pdf_to_word.docx')
    #pdf_to_pptx('media/pdf_to_word.docx', 'media/pdf_to_pptx.pptx')
    #pdf_to_excel(pdf_file, 'media/pdf_to_excel.xlsx')
    #word_to_pdf('media/pdf_to_word.docx', 'media/word_to_pdf.pdf')
    #ppt_to_pdf('media/pdf_to_pptx.pptx', 'media/ppt_to_pdf.pdf')
    #excel_to_pdf('media/pdf_to_excel.xlsx', 'media/excel_to_pdf.pdf')
    #pdf_to_jpg(pdf_file, 'media/pdf_to_jpg')
    #jpg_to_pdf('media/pdf_to_jpg', 'media/jpg_to_pdf.pdf')
    #html_to_pdf('https://www.facebook.com/', 'media/html_to_pdf.pdf')
    #add_page_numbers(pdf_file, 'media/add_page_numbers.pdf', (0,0))
    #protect_pdf(pdf_file, 'media/protected.pdf', '1111')
    #unlock_pdf("media/protected.pdf", "media/unlocked.pdf", "1111")

    parser = argparse.ArgumentParser(description='PDF Utility Script')
    parser.add_argument('function', choices=['merge', 'split', 'compress', 'pdf_to_word', 'pdf_to_pptx',
                                             'pdf_to_excel', 'word_to_pdf', 'ppt_to_pdf', 'excel_to_pdf', 'pdf_to_jpg',
                                             'jpg_to_pdf', 'html_to_pdf', 'add_page_numbers', 'protect_pdf',
                                             'unlock_pdf'], help='Choose a function to execute')
    parser.add_argument('--input', help='Input file')
    parser.add_argument('--output', help='Output file')
    parser.add_argument('--start_page', type=int, help='Start page (for split operation)')
    parser.add_argument('--end_page', type=int, help='End page (for split operation)')
    parser.add_argument('--dpi', type=int, help='DPI (for compress operation)')
    parser.add_argument('--password', help='Password (for protect_pdf and unlock_pdf operations)')
    parser.add_argument('--url', help='URL (for html_to_pdf operation)')
    parser.add_argument('--position', choices=['left_bottom', 'left_middle', 'left_top', 'middle_top',
                                               'middle_middle', 'middle_bottom', 'right_bottom', 'right_middle',
                                               'right_top'], help='Position for page numbers')
    args = parser.parse_args()

    if args.function == 'merge':
        merge_pdfs(args.input.split(','), args.output)
    elif args.function == 'split':
        split_pdf(args.input, args.start_page, args.end_page, args.output)
    elif args.function == 'compress':
        compress_pdf(args.input, args.output, args.dpi)
    elif args.function == 'pdf_to_word':
        pdf_to_word(args.input, args.output)
    elif args.function == 'pdf_to_pptx':
        pdf_to_pptx(args.input, args.output)
    elif args.function == 'pdf_to_excel':
        pdf_to_excel(args.input, args.output)
    elif args.function == 'word_to_pdf':
        word_to_pdf(args.input, args.output)
    elif args.function == 'ppt_to_pdf':
        ppt_to_pdf(args.input, args.output)
    elif args.function == 'excel_to_pdf':
        excel_to_pdf(args.input, args.output)
    elif args.function == 'pdf_to_jpg':
        pdf_to_jpg(args.input, args.output)
    elif args.function == 'jpg_to_pdf':
        jpg_to_pdf(args.input, args.output)
    elif args.function == 'html_to_pdf':
        html_to_pdf(args.url, args.output)
    elif args.function == 'add_page_numbers':
        positions = {
            'left_bottom': (10, 10),
            'left_middle': (10, 148),
            'left_top': (10, 280),
            'middle_top': (105, 280),
            'middle_middle': (105, 148),
            'middle_bottom': (105, 10),
            'right_bottom': (200, 10),
            'right_middle': (200, 148),
            'right_top': (200, 280)
        }
        add_page_numbers(args.input, args.output, positions[args.position])
    elif args.function == 'protect_pdf':
        protect_pdf(args.input, args.output, args.password)
    elif args.function == 'unlock_pdf':
        unlock_pdf(args.input, args.output, args.password)
